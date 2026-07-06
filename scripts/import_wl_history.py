"""One-time import of historical win/loss data from the legacy W-L tracker
spreadsheet into league_history/league_seasons/matchups.

The source workbook lives outside this repo (a personal spreadsheet, never
committed) — pass its path on the command line. Two sheets per year feed the
import: `<year> Results` (flat Year/Week/League/Outcome/PF/PA/Differential/
Playoff rows, one per week) populates `matchups`; `Aggregated Data` (thirteen
fixed 10-row blocks, most recent year first, one row per league, zero-padded
for leagues that didn't exist that year) populates `league_seasons`.

`league_history` rows are created as needed (matched by name) and marked
`active` based on whether that league appears anywhere in the most recent
year's data — this only runs once against the full 2013-2025 history, so
"most recent" is fixed at 2025 rather than computed from today's date.

Idempotent: safe to rerun (upserts on each table's natural key).

Usage:
    python -m scripts.import_wl_history "C:\\path\\to\\W-L 2025.xlsx"
"""

import argparse
import sqlite3
from pathlib import Path

import openpyxl

from ffassistant.db import get_connection

FIRST_YEAR = 2013
LATEST_YEAR = 2025
AGG_FIRST_ROW = 2
AGG_BLOCK_SIZE = 10

PLAYOFF_ROUND_MAP = {
    "Semifinal": "semifinal",
    "Final": "final",
    "3rd": "third_place",
}


def import_wl_history(
    xlsx_path: Path,
    conn: sqlite3.Connection | None = None,
    first_year: int = FIRST_YEAR,
    latest_year: int = LATEST_YEAR,
) -> dict:
    conn = conn or get_connection()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        league_names = _collect_all_league_names(wb, first_year, latest_year)
        active_names = _collect_league_names_for_year(wb, latest_year, latest_year)
        history_ids = _upsert_league_history(conn, league_names, active_names)

        matchup_count = 0
        season_count = 0
        for year in range(first_year, latest_year + 1):
            matchup_count += _import_matchups(conn, wb, year, history_ids)
            season_count += _import_league_seasons(conn, wb, year, history_ids, latest_year)

        conn.commit()
        return {"leagues": len(history_ids), "matchups": matchup_count, "league_seasons": season_count}
    finally:
        wb.close()


def _agg_block_bounds(year: int, latest_year: int) -> tuple[int, int]:
    block_index = latest_year - year
    start = AGG_FIRST_ROW + AGG_BLOCK_SIZE * block_index
    return start, start + AGG_BLOCK_SIZE - 1


def _collect_league_names_for_year(wb, year: int, latest_year: int) -> set:
    names = set()

    ws = wb[f"{year} Results"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        league = row[2]
        if league:
            names.add(league)

    start, end = _agg_block_bounds(year, latest_year)
    ws = wb["Aggregated Data"]
    for r in range(start, end + 1):
        name = ws.cell(row=r, column=1).value
        if name and name != 0:
            names.add(name)

    return names


def _collect_all_league_names(wb, first_year: int, latest_year: int) -> set:
    names = set()
    for year in range(first_year, latest_year + 1):
        names |= _collect_league_names_for_year(wb, year, latest_year)
    return names


def _upsert_league_history(conn, league_names: set, active_names: set) -> dict:
    history_ids = {}
    for name in sorted(league_names):
        is_active = 1 if name in active_names else 0
        row = conn.execute(
            """
            INSERT INTO league_history (name, active) VALUES (?, ?)
            ON CONFLICT (name) DO UPDATE SET active = excluded.active
            RETURNING league_history_id
            """,
            (name, is_active),
        ).fetchone()
        history_ids[name] = row["league_history_id"]
    return history_ids


RAW_OUTCOME_MAP = {"W": "W", "L": "L", "D": "T"}


def _derive_outcome(outcome, points_for, points_against):
    """A handful of legacy rows have real scores but a blank Outcome cell
    (manual-entry gaps) — recompute it from the scores rather than dropping
    real data. An explicit W/L/D is trusted as-is (mapped to W/L/T)."""
    if outcome is None:
        if points_for > points_against:
            return "W"
        if points_for < points_against:
            return "L"
        return "T"
    return RAW_OUTCOME_MAP.get(outcome)


def _import_matchups(conn, wb, year: int, history_ids: dict) -> int:
    ws = wb[f"{year} Results"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_year, week, league, outcome, points_for, points_against, _differential, playoff = row
        if league is None or week is None or points_for is None:
            continue  # not yet played/recorded — no real data to import for this week
        outcome = _derive_outcome(outcome, points_for, points_against)
        if outcome is None:
            # Guillotine (survivor-elimination format) records a weekly rank
            # here instead of a head-to-head result — no real W/L to import.
            continue
        conn.execute(
            """
            INSERT INTO matchups
                (league_history_id, season, week, points_for, points_against, outcome, playoff_round)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (league_history_id, season, week) DO UPDATE SET
                points_for = excluded.points_for,
                points_against = excluded.points_against,
                outcome = excluded.outcome,
                playoff_round = excluded.playoff_round
            """,
            (history_ids[league], row_year, week, points_for, points_against, outcome, PLAYOFF_ROUND_MAP.get(playoff)),
        )
        count += 1
    return count


def _import_league_seasons(conn, wb, year: int, history_ids: dict, latest_year: int) -> int:
    start, end = _agg_block_bounds(year, latest_year)
    ws = wb["Aggregated Data"]
    count = 0
    for r in range(start, end + 1):
        name = ws.cell(row=r, column=1).value
        if not name or name == 0:
            continue
        wins, losses, ties, place, buy_in, max_payout, won = (ws.cell(row=r, column=c).value for c in range(2, 9))
        conn.execute(
            """
            INSERT INTO league_seasons
                (league_history_id, season, wins, losses, ties, buy_in, max_payout, actual_payout, finish_position)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (league_history_id, season) DO UPDATE SET
                wins = excluded.wins,
                losses = excluded.losses,
                ties = excluded.ties,
                buy_in = excluded.buy_in,
                max_payout = excluded.max_payout,
                actual_payout = excluded.actual_payout,
                finish_position = excluded.finish_position
            """,
            (history_ids[name], year, wins, losses, ties, buy_in, max_payout, won, place),
        )
        count += 1
    return count


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", help="Path to the legacy W-L tracker spreadsheet")
    args = parser.parse_args()

    result = import_wl_history(Path(args.xlsx_path))
    print(
        f"Imported {result['leagues']} leagues, {result['matchups']} matchup rows, "
        f"{result['league_seasons']} league-season rows"
    )
