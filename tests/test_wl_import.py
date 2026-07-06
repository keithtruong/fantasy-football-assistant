import tempfile
import unittest
from pathlib import Path

import openpyxl

from scripts.import_wl_history import import_wl_history
from tests.test_import_scripts import make_conn


def _build_fixture_workbook(path: Path) -> None:
    """Two-year fixture (2024-2025) mirroring the real workbook's shape:
    `<year> Results` (Year, Week, League, Outcome, PF, PA, Differential,
    Playoff) and a 2-block `Aggregated Data` sheet (2025's block first, then
    2024's — matching the real file's most-recent-year-first ordering).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    results_2025 = wb.create_sheet("2025 Results")
    results_2025.append(["Year", "Week", "League", "Outcome", "PF", "PA", "Differential", "Playoff"])
    results_2025.append([2025, 1, "Alpha", "W", 120.5, 100.0, 20.5, None])
    results_2025.append([2025, 1, "Beta", "L", 90.0, 95.0, -5.0, None])
    # Blank Outcome cell but real scores -> outcome should be derived (Alpha week 2 = W).
    results_2025.append([2025, 2, "Alpha", None, 110.0, 108.0, 2.0, None])
    # Survivor-style non-W/L "outcome" (a weekly rank) with PF==PA -> must be skipped entirely.
    results_2025.append([2025, 1, "Gamma", 5, 88.0, 88.0, 0, None])
    # Not yet played (no scores at all) -> must be skipped entirely.
    results_2025.append([2025, 3, "Alpha", None, None, None, 0, None])
    results_2025.append([2025, 16, "Alpha", "W", 130.0, 90.0, 40.0, "Semifinal"])

    results_2024 = wb.create_sheet("2024 Results")
    results_2024.append(["Year", "Week", "League", "Outcome", "PF", "PA", "Differential", "Playoff"])
    results_2024.append([2024, 1, "Alpha", "L", 80.0, 100.0, -20.0, None])
    # A tie, spelled "D" in the legacy sheet -> should map to "T".
    results_2024.append([2024, 2, "Alpha", "D", 100.0, 100.0, 0, None])
    # Delta only ever appears in 2024 -> should come back inactive.
    results_2024.append([2024, 1, "Delta", "W", 100.0, 90.0, 10.0, None])

    agg = wb.create_sheet("Aggregated Data")
    agg.append(["Team", "W", "L", "T", "Place", "Buy-in", "Max", "Won"])
    # Blocks are a fixed 10 rows each in the real workbook, most-recent-year
    # first — the import's block-offset math depends on that fixed size, so
    # the fixture must pad to 10 rows per block just like the real file does.
    # 2025 block (rows 2-11): Alpha, Beta, Gamma present, rest zero-padded.
    agg.append(["Alpha", 10, 5, 0, 1, 50, 500, 500])
    agg.append(["Beta", 8, 7, 0, 3, 20, 100, 0])
    agg.append(["Gamma", 0, 0, 0, 8, 50, 900, 0])
    for _ in range(7):
        agg.append([0, 0, 0, 0, 0, 0, 0, 0])
    # 2024 block (rows 12-21): Alpha and Delta present, rest zero-padded.
    agg.append(["Alpha", 9, 6, 0, 2, 50, 500, 0])
    agg.append(["Delta", 4, 3, 0, 5, 0, 0, 0])
    for _ in range(8):
        agg.append([0, 0, 0, 0, 0, 0, 0, 0])

    wb.save(path)


class TestImportWlHistory(unittest.TestCase):
    def setUp(self):
        self.conn = make_conn()
        # ignore_cleanup_errors: openpyxl .xlsx files can get a transient lock
        # from Windows antivirus scanning right after being written/read —
        # unrelated to the import logic itself.
        self._tmpdir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.xlsx_path = Path(self._tmpdir.name) / "fixture.xlsx"
        _build_fixture_workbook(self.xlsx_path)

    def tearDown(self):
        self._tmpdir.cleanup()

    def _import(self):
        return import_wl_history(self.xlsx_path, conn=self.conn, first_year=2024, latest_year=2025)

    def test_league_history_created_with_active_derived_from_latest_year(self):
        self._import()
        rows = {r["name"]: r["active"] for r in self.conn.execute("SELECT name, active FROM league_history")}
        self.assertEqual(rows, {"Alpha": 1, "Beta": 1, "Gamma": 1, "Delta": 0})

    def test_matchup_rows_imported_with_correct_fields(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT m.* FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Alpha' AND m.season = 2025 AND m.week = 1
            """
        ).fetchone()
        self.assertEqual(row["points_for"], 120.5)
        self.assertEqual(row["points_against"], 100.0)
        self.assertEqual(row["outcome"], "W")

    def test_blank_outcome_derived_from_scores(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT m.outcome FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Alpha' AND m.season = 2025 AND m.week = 2
            """
        ).fetchone()
        self.assertEqual(row["outcome"], "W")

    def test_raw_d_outcome_maps_to_tie(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT m.outcome FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Alpha' AND m.season = 2024 AND m.week = 2
            """
        ).fetchone()
        self.assertEqual(row["outcome"], "T")

    def test_survivor_style_numeric_outcome_skipped(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT COUNT(*) AS c FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Gamma' AND m.season = 2025 AND m.week = 1
            """
        ).fetchone()
        self.assertEqual(row["c"], 0)

    def test_unplayed_week_skipped(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT COUNT(*) AS c FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Alpha' AND m.season = 2025 AND m.week = 3
            """
        ).fetchone()
        self.assertEqual(row["c"], 0)

    def test_playoff_round_mapped(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT m.playoff_round FROM matchups m JOIN league_history lh ON lh.league_history_id = m.league_history_id
            WHERE lh.name = 'Alpha' AND m.season = 2025 AND m.week = 16
            """
        ).fetchone()
        self.assertEqual(row["playoff_round"], "semifinal")

    def test_league_seasons_block_offset_lands_on_correct_year(self):
        self._import()
        rows = {
            r["season"]: (r["wins"], r["losses"], r["finish_position"])
            for r in self.conn.execute(
                """
                SELECT ls.season, ls.wins, ls.losses, ls.finish_position
                FROM league_seasons ls JOIN league_history lh ON lh.league_history_id = ls.league_history_id
                WHERE lh.name = 'Alpha'
                """
            )
        }
        self.assertEqual(rows[2025], (10, 5, 1))
        self.assertEqual(rows[2024], (9, 6, 2))

    def test_zero_padded_rows_skipped(self):
        self._import()
        row = self.conn.execute(
            """
            SELECT COUNT(*) AS c FROM league_seasons ls JOIN league_history lh ON lh.league_history_id = ls.league_history_id
            WHERE lh.name = 'Beta' AND ls.season = 2024
            """
        ).fetchone()
        self.assertEqual(row["c"], 0)

    def test_rerun_is_idempotent(self):
        first = self._import()
        second = self._import()
        self.assertEqual(first, second)
        total_matchups = self.conn.execute("SELECT COUNT(*) AS c FROM matchups").fetchone()["c"]
        self.assertEqual(total_matchups, first["matchups"])


if __name__ == "__main__":
    unittest.main()
